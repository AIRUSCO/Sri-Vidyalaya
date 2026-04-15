from openpyxl import Workbook
from datetime import date

wb = Workbook()
ws = wb.active
ws.title = "Launch Tracker"

headers = ["Category", "Task", "Status", "Priority", "Target", "Notes"]
ws.append(headers)

completed = [
    ["Completed", "Updated social media links across all website pages", "Done", "High", "All pages", "Synced WhatsApp, Instagram, Facebook, LinkedIn links"],
    ["Completed", "Removed legacy L3 files and standardized filenames", "Done", "High", "File structure", "Deleted index-L3/about-L3/style-L3 and created clean what-we-stand-for page"],
    ["Completed", "Standardized header and footer across all pages", "Done", "High", "Global layout", "Unified navigation/footer sections for consistency"],
    ["Completed", "Fixed broken image paths on About and What We Stand For pages", "Done", "High", "about.html, what-we-stand-for.html", "Corrected image filenames and paths"],
    ["Completed", "Adjusted image/text alignment for About and What We Stand For sections", "Done", "Medium", "About + Values sections", "Improved equal-height behavior and responsive layout"],
    ["Completed", "Created new KV Greens Services page", "Done", "High", "services.html", "Added full service sections with alternating text/image layout"],
    ["Completed", "Enhanced services page spacing, button consistency, and hero description", "Done", "Medium", "services.html", "Applied service-section spacing and unified Learn More button style"],
]

pending = [
    ["Pending", "Create missing service detail pages linked from site", "To Do", "High", "/services/*.html", "Currently linked in multiple pages but files are missing (forestry, land-management, sustainable-agriculture, etc.)"],
    ["Pending", "Review and finalize all content copy (proofread + brand voice)", "To Do", "Medium", "All pages", "Check headings, descriptions, and consistency"],
    ["Pending", "Cross-browser and mobile QA", "To Do", "High", "Chrome, Edge, Safari, mobile sizes", "Validate layout, spacing, and navigation behavior"],
    ["Pending", "Validate all internal/external links and form submission flow", "To Do", "High", "Full website", "Test contact form success path and all CTA links"],
    ["Pending", "SEO launch checklist", "To Do", "Medium", "All pages", "Meta titles/descriptions, OG tags, sitemap.xml, robots.txt, canonical tags"],
    ["Pending", "Performance optimization", "To Do", "Medium", "Images + scripts", "Compress heavy images, lazy-load where needed, review unused assets"],
    ["Pending", "Accessibility check", "To Do", "Medium", "All pages", "Alt text quality, keyboard navigation, color contrast, heading hierarchy"],
    ["Pending", "Production deployment and post-launch smoke test", "To Do", "High", "Hosting + domain", "Deploy, clear cache/CDN, verify SSL and final URLs"],
]

for row in completed + pending:
    ws.append(row)

# simple formatting
from openpyxl.styles import Font, PatternFill, Alignment
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

for col, width in {"A":12, "B":68, "C":12, "D":10, "E":28, "F":80}.items():
    ws.column_dimensions[col].width = width

for r in range(2, ws.max_row + 1):
    ws.cell(r, 1).alignment = Alignment(vertical="top", wrap_text=True)
    ws.cell(r, 2).alignment = Alignment(vertical="top", wrap_text=True)
    ws.cell(r, 6).alignment = Alignment(vertical="top", wrap_text=True)

file_name = "KV-Greens-Launch-Tracker.xlsx"
wb.save(file_name)
print(file_name)
